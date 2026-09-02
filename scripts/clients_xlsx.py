#!/usr/bin/env python3
"""Excel 客戶資料：一 sheet 一間公司 → clients/{slug}/client.json

  python3 scripts/clients_xlsx.py init
  python3 scripts/clients_xlsx.py export
  python3 scripts/clients_xlsx.py sync
  python3 scripts/clients_xlsx.py pack
  python3 scripts/clients_xlsx.py preview example-pd
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    sys.stderr.write("需要 openpyxl：pip install -r scripts/requirements.txt\n")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "clients.xlsx"
CLIENTS_DIR = ROOT / "clients"
DATA_JS = ROOT / "js" / "data.js"

SLUG_RE = re.compile(r"^[a-z][a-z0-9-]{0,30}$")
SKIP_PREFIX = "_"

META_LABEL = "clientName"
QR_LABEL = "qrCode"
HEADER_ROW = 4
DATA_START_ROW = 5

COLUMNS = [
    ("name", "APP 名稱", 22),
    ("version", "版本", 12),
    ("requirement", "系統需求", 36),
    ("tagline", "標題說明", 42),
    ("feature_left", "左字卡", 42),
    ("feature_right_top", "右上字卡", 42),
    ("feature_right_bottom", "右下字卡", 36),
    ("icon", "圖示檔名", 28),
    ("screenshot_1", "截圖 1（必填）", 36),
    ("screenshot_2", "截圖 2（有＝雙機）", 36),
    ("ios", "App Store 網址", 36),
    ("google_play", "Google Play 網址", 36),
    ("android", "Android APK 網址", 32),
    ("user_guide", "使用說明（網址或檔名）", 28),
]

COL_KEYS = [key for key, _, _ in COLUMNS]
FEATURE_MAP = {
    "feature_left": "left",
    "feature_right_top": "rightTop",
    "feature_right_bottom": "rightBottom",
}
FILE_FIELDS = ("icon", "screenshot_1", "screenshot_2", "user_guide")
FOLDER_UNSAFE = re.compile(r'[\\/:*?"<>|]+')
CANON_STEMS = {
    "icon": "icon",
    "screenshot_1": "screenshot-1",
    "screenshot_2": "screenshot-2",
    "user_guide": "guide",
}
QR_STEM = "qrcode"
PACK_SKIP_SLUGS = {"example-pd", "example-sparse", "zoe"}
PACK_FILES = ("index.html",)
PACK_DIRS = ("css", "js", "img")
PACK_ZIP = ROOT / "dist" / "csi-appstore.zip"

BLUE = "0D63BA"
LIGHT = "F6F7F9"
GRAY = "22354A"
HINT = "6B7A8D"

FILL_HEAD = PatternFill("solid", fgColor=BLUE)
FILL_META = PatternFill("solid", fgColor="D0DEF3")
FILL_HINT = PatternFill("solid", fgColor=LIGHT)
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_META = Font(name="Calibri", bold=True, color=BLUE, size=12)
FONT_HINT = Font(name="Calibri", italic=True, color=HINT, size=10)
FONT_BODY = Font(name="Calibri", color=GRAY, size=11)
THIN = Border(
    left=Side(style="thin", color="C6CDD8"),
    right=Side(style="thin", color="C6CDD8"),
    top=Side(style="thin", color="C6CDD8"),
    bottom=Side(style="thin", color="C6CDD8"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


MD_LINK = re.compile(r"^\[([^\]]+)\]\(([^)]+)\)$")


def unwrap_href(raw: str) -> str:
    text = raw.strip()
    if not text:
        return ""
    match = MD_LINK.match(text)
    if match:
        return match.group(2).strip()
    return text


def cell_href(cell) -> str:
    text = unwrap_href(cell_str(cell.value))
    if cell.hyperlink and cell.hyperlink.target:
        target = unwrap_href(str(cell.hyperlink.target))
        if text and "://" not in text and not text.startswith("#"):
            return text
        return target
    return text


def is_client_sheet(name: str) -> bool:
    return bool(name) and not name.startswith(SKIP_PREFIX)


def name_to_slug(name: str) -> str:
    slug = name.strip().lower()
    slug = FOLDER_UNSAFE.sub("-", slug)
    slug = re.sub(r"[\s_]+", "-", slug)
    slug = re.sub(r"-+", "-", slug).strip("-")
    return slug[:31]


def legacy_product_folder(name: str) -> str:
    folder = FOLDER_UNSAFE.sub("-", name).strip()
    folder = re.sub(r"\s+", " ", folder)
    return folder


def validate_app_slug(client_slug: str, display_name: str, app_slug: str) -> str | None:
    if not app_slug:
        return f"{client_slug}「{display_name}」無法從 APP 名稱產生資料夾名"
    if not SLUG_RE.match(app_slug):
        return (
            f"{client_slug}「{display_name}」產生的資料夾名「{app_slug}」格式不對。"
            "請調整 APP 名稱（需能轉成小寫英文、數字、連字號）。"
        )
    return None


def is_remote_or_shared(value: str) -> bool:
    return (not value) or value.startswith("#") or "://" in value or value.startswith("img/")


def localize_qr(slug: str, value: str) -> str:
    if is_remote_or_shared(value):
        return value
    prefix = f"clients/{slug}/"
    if value.startswith("clients/") and not value.startswith(prefix):
        return value
    return f"{prefix}{Path(value).name}"


def localize_asset(slug: str, app_slug: str, value: str) -> str:
    if is_remote_or_shared(value):
        return value
    prefix = f"clients/{slug}/"
    if value.startswith("clients/") and not value.startswith(prefix):
        return value
    return f"{prefix}{app_slug}/{Path(value).name}"


def validate_slug(name: str) -> str | None:
    if not is_client_sheet(name):
        return None
    if not SLUG_RE.match(name):
        return (
            f"工作表「{name}」不能當網址短名。"
            "請用小寫英文、數字、連字號，最長 31 字，例如 taipei-pd。"
        )
    return None


def apply_column_widths(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 28
    for i, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def write_sheet(
    ws: Worksheet, client_name: str, rows: list[dict], example: bool, qr_code: str = ""
) -> None:
    ws["A1"] = META_LABEL
    ws["B1"] = client_name
    ws["C1"] = "整頁只填一次 → 左上角顯示這格文字"
    ws["A1"].font = FONT_META
    ws["B1"].font = Font(name="Calibri", bold=True, size=14, color=GRAY)
    ws["C1"].font = FONT_HINT
    ws["A1"].fill = FILL_META
    ws["B1"].fill = FILL_META
    ws.merge_cells("B1:E1")

    ws["A2"] = QR_LABEL
    ws["B2"] = qr_code
    ws["C2"] = "← 填這格左邊（B2）"
    ws["A2"].font = FONT_META
    ws["B2"].font = FONT_BODY
    ws["C2"].font = FONT_HINT
    ws["A2"].fill = FILL_META
    ws["B2"].fill = FILL_META
    ws["B2"].comment = Comment(
        "頁尾 QR 圖片：填網址或檔名。沒填就不顯示。",
        "CSI",
    )

    for i, (_, zh, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(3, i, zh)
        cell.font = FONT_HINT
        cell.fill = FILL_HINT
        cell.alignment = WRAP

    for i, (key, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, i, key)
        cell.font = FONT_WHITE
        cell.fill = FILL_HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

    tagline = ws.cell(HEADER_ROW, COL_KEYS.index("tagline") + 1)
    tagline.comment = Comment(
        "APP 名稱下方的說明。選填；空白或 # 就不顯示。與左／右字卡無關。",
        "CSI",
    )
    shot2 = ws.cell(HEADER_ROW, COL_KEYS.index("screenshot_2") + 1)
    shot2.comment = Comment(
        "留空＝版型二（單機）。有檔名＝版型一（雙機）。列表順序不切版型。",
        "CSI",
    )

    extra = 4 if example else 8
    for r in range(DATA_START_ROW, DATA_START_ROW + len(rows) + extra):
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(r, c, "")
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for offset, row in enumerate(rows):
        r = DATA_START_ROW + offset
        for c, key in enumerate(COL_KEYS, start=1):
            ws.cell(r, c, row.get(key, ""))

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{HEADER_ROW}"
    apply_column_widths(ws)
    ws.sheet_view.showGridLines = False


def write_readme(ws: Worksheet) -> None:
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws["A1"] = "怎麼用這本 Excel"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=BLUE)
    lines = [
        "",
        "一本活頁簿管全部客戶。一個工作表 = 一間公司。",
        "工作表名稱就是網址短名（小寫、連字號，例如 hillsdale），也是 ?client= 參數。",
        "短名一旦給客戶就不要改。_ 開頭的表（_readme、_template）不會匯出、不會上線。",
        "",
        "新客戶",
        "1. 複製 _template，把工作表改名成短名",
        "2. B1 填客戶全名（左上角）；B2 填頁尾 QR（檔名或網址；沒填不顯示）",
        "3. 從第 5 列起，一列一款 APP",
        "4. 圖與說明書放到 clients/短名/（由 APP 名稱自動產生的資料夾）/ ；QR 放 clients/短名/",
        "5. Excel 只填正確檔名，不要把圖貼進儲存格",
        "6. 存檔後執行：python3 scripts/clients_xlsx.py sync",
        "",
        "改完資料一律 sync（不要只跑 export，否則檔名／資料夾不會對）",
        "本機預覽：python3 -m http.server 5501  →  http://127.0.0.1:5501/?client=短名",
        "上線包：python3 scripts/clients_xlsx.py pack  →  dist/csi-appstore.zip",
        "",
        "檔名（必須用這套，不要自創）",
        "整頁 QR          qrcode.（png/jpg…）     放在 clients/短名/",
        "APP 圖示         icon.（ext）             放在該款 APP 資料夾",
        "截圖 1           screenshot-1.（ext）     同上（必填）",
        "截圖 2           screenshot-2.（ext）     有才放；有＝雙機、沒有＝單機",
        "使用說明檔       guide.（ext）            有才放",
        "產品資料夾名稱由 APP 名稱自動產生（小寫、連字號，例如 Active Response → active-response）。",
        "",
        "空白與 #",
        "沒填或填 # 都當沒有：不寫進 JSON，頁面不出現對應按鈕／說明／字卡。",
        "",
        "每一款 APP 的欄（第 3 列中文、第 4 列英文；不要改英文表頭）",
        "name                 APP 名稱（頁面顯示；sync 會自動轉成資料夾名）",
        "version              版本（截圖下方，與系統需求同一行）",
        "requirement          系統需求",
        "tagline              標題說明：名稱正下方那一段。選填。",
        "                     與字卡無關，不要把左／右字卡的句子抄來這裡。",
        "                     儲存格換行，頁面上會變成多段。",
        "feature_left         左字卡（寬螢幕左邊卡片；手機勾勾清單）",
        "feature_right_top    右上字卡",
        "feature_right_bottom 右下字卡（只出現在寬螢幕右下與手機清單，不上標題下）",
        "icon / screenshot_1 / screenshot_2 / ios / google_play / android / user_guide",
        "                     有填才顯示圖示、截圖、商店按鈕、Download User Guide",
        "",
        "版型",
        "screenshot_1 必填。screenshot_2 留空＝單機；有填＝雙機。列表第幾款不影響版型。",
        "同一頁 RWD：寬螢幕 Desktop（截圖在上、下載按鈕在下），平板與手機用 Mobile。",
        "這兩種排法已鎖定，不要為了「下載鈕比較上面」去改 Desktop。",
        "",
        "目前客戶工作表",
        "smart-industry-center、harbor-county-sheriff、critical-technology、hillsdale",
        "",
        "不要手改 clients/短名/client.json，一律改 Excel 再 sync。",
    ]
    for i, line in enumerate(lines, start=2):
        ws[f"A{i}"] = line
        ws[f"A{i}"].font = FONT_BODY
    ws.column_dimensions["A"].width = 92
    ws.sheet_view.showGridLines = False


EXAMPLE_PD = {
    "slug": "example-pd",
    "clientName": "Client Name",
    "rows": [
        {
            "name": "Mobile MDT",
            "version": "V2.4.1",
            "requirement": "Android 4.3 / iOS 11.0 or above",
            "tagline": "In-car records, maps, and dispatch in one place.",
            "feature_left": "Arrive Informed. Respond Faster. Save Lives.",
            "feature_right_top": "From Floor Plans to RMS: Total Real-Time Command.",
            "feature_right_bottom": "Zero Delay. Total Clarity.",
            "icon": "img/apps/icon-InfoMDT.png",
            "screenshot_1": "img/ui/screenshot-template.png",
            "screenshot_2": "img/ui/screenshot-template.png",
            "ios": "https://apps.apple.com/example",
            "google_play": "https://play.google.com/example",
            "android": "https://example.com/mdt.apk",
            "user_guide": "",
        },
        {
            "name": "Active Response",
            "version": "V1.8.0",
            "requirement": "Android 4.3 / iOS 11.0 or above",
            "tagline": "",
            "feature_left": "Manage cases everywhere you want.",
            "feature_right_top": "Arrive Informed. Respond Faster. Save Lives.",
            "feature_right_bottom": "Zero Delay. Total Clarity.",
            "icon": "img/apps/icon-ActiveResponse.png",
            "screenshot_1": "img/ui/screenshot-template.png",
            "screenshot_2": "",
            "ios": "https://apps.apple.com/example",
            "google_play": "https://play.google.com/example",
            "android": "",
            "user_guide": "",
        },
    ],
}

EXAMPLE_SPARSE = {
    "slug": "example-sparse",
    "clientName": "Harbor County Sheriff",
    "rows": [
        {
            "name": "Field Notes",
            "version": "V1.0.0",
            "requirement": "iOS 14.0 or above",
            "tagline": "",
            "feature_left": "Capture the scene before you leave it.",
            "feature_right_top": "",
            "feature_right_bottom": "",
            "icon": "img/apps/icon-InfoMDT.png",
            "screenshot_1": "img/ui/screenshot-template.png",
            "screenshot_2": "",
            "ios": "https://apps.apple.com/example",
            "google_play": "",
            "android": "",
            "user_guide": "",
        }
    ],
}


def build_workbook() -> Workbook:
    wb = Workbook()
    readme = wb.active
    readme.title = "_readme"
    write_readme(readme)

    template = wb.create_sheet("_template")
    write_sheet(template, "", [], example=False)

    pd = wb.create_sheet(EXAMPLE_PD["slug"])
    write_sheet(
        pd,
        EXAMPLE_PD["clientName"],
        EXAMPLE_PD["rows"],
        example=True,
        qr_code="img/QRcode.png",
    )

    sparse = wb.create_sheet(EXAMPLE_SPARSE["slug"])
    write_sheet(
        sparse,
        EXAMPLE_SPARSE["clientName"],
        EXAMPLE_SPARSE["rows"],
        example=True,
        qr_code="img/QRcode.png",
    )
    return wb


def cmd_init(force: bool) -> int:
    if XLSX_PATH.exists() and not force:
        sys.stderr.write(
            f"已有 {XLSX_PATH.name}。若要重建範例檔，請先備份再加 --force。\n"
        )
        return 1
    build_workbook().save(XLSX_PATH)
    print(f"已寫入 {XLSX_PATH.relative_to(ROOT)}")
    print("複製 _template、把工作表改成短名，填完後執行 export。")
    return 0


def find_header_row(ws: Worksheet) -> int | None:
    for r in range(1, 12):
        values = [cell_str(ws.cell(r, c).value) for c in range(1, 20)]
        if "name" in values and "screenshot_1" in values:
            return r
    return None


def read_meta(ws: Worksheet, header_row: int, label: str) -> str:
    for r in range(1, header_row):
        if cell_str(ws.cell(r, 1).value) == label:
            return cell_href(ws.cell(r, 2))
    return ""


def col_index_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    found = {}
    for c in range(1, 40):
        key = cell_str(ws.cell(header_row, c).value)
        if key:
            found[key] = c
    return found


OBSOLETE_COLUMNS = ("slug",)


def remove_obsolete_columns(ws: Worksheet) -> bool:
    header_row = find_header_row(ws)
    if not header_row:
        return False
    changed = False
    while True:
        cols = col_index_map(ws, header_row)
        removed = False
        for key in OBSOLETE_COLUMNS:
            c = cols.get(key)
            if not c:
                continue
            ws.delete_cols(c)
            changed = True
            removed = True
            break
        if not removed:
            break
    if changed:
        last = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A{header_row}:{last}{header_row}"
    return changed


def ensure_columns(ws: Worksheet) -> bool:
    header_row = find_header_row(ws)
    if not header_row:
        return False
    changed = False
    for i, (key, zh, width) in enumerate(COLUMNS, start=1):
        if key in col_index_map(ws, header_row):
            continue
        ws.insert_cols(i)
        hint_row = header_row - 1 if header_row > 1 else header_row
        zh_cell = ws.cell(hint_row, i, zh)
        zh_cell.font = FONT_HINT
        zh_cell.fill = FILL_HINT
        zh_cell.alignment = WRAP
        key_cell = ws.cell(header_row, i, key)
        key_cell.font = FONT_WHITE
        key_cell.fill = FILL_HEAD
        key_cell.alignment = Alignment(horizontal="center", vertical="center")
        key_cell.border = THIN
        if key == "tagline":
            key_cell.comment = Comment(
                "APP 名稱下方的說明。選填；空白或 # 就不顯示。與左／右字卡無關。",
                "CSI",
            )
        ws.column_dimensions[get_column_letter(i)].width = width
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            cell = ws.cell(r, i)
            cell.font = FONT_BODY
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        changed = True
    last = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A{header_row}:{last}{header_row}"
    return changed


def clear_stale_file_hyperlinks(ws: Worksheet) -> bool:
    header_row = find_header_row(ws)
    if not header_row:
        return False
    cols = col_index_map(ws, header_row)
    changed = False
    for field in FILE_FIELDS:
        c = cols.get(field)
        if not c:
            continue
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            cell = ws.cell(r, c)
            text = cell_str(cell.value)
            if cell.hyperlink and text and "://" not in text:
                cell.hyperlink = None
                changed = True
    return changed


def ensure_book_columns(wb) -> bool:
    changed = False
    for name in wb.sheetnames:
        if name == "_readme":
            continue
        if remove_obsolete_columns(wb[name]):
            changed = True
        if ensure_columns(wb[name]):
            changed = True
        if clear_stale_file_hyperlinks(wb[name]):
            changed = True
    return changed


def row_to_app(ws: Worksheet, r: int, cols: dict[str, int], slug: str) -> dict | None:
    def get(key: str) -> str:
        c = cols.get(key)
        if not c:
            return ""
        return cell_href(ws.cell(r, c))

    name = get("name")
    if not name:
        return None

    app_slug = name_to_slug(name)

    features = {}
    for col_key, json_key in FEATURE_MAP.items():
        text = get(col_key)
        if text:
            features[json_key] = text

    shots = [
        localize_asset(slug, app_slug, s)
        for s in (get("screenshot_1"), get("screenshot_2"))
        if s
    ]

    app = {"name": name}
    version = get("version")
    requirement = get("requirement")
    icon = get("icon")
    if version:
        app["version"] = version
    if requirement:
        app["requirement"] = requirement
    tagline = get("tagline")
    if tagline and tagline != "#":
        app["tagline"] = tagline
    if features:
        app["features"] = features
    if icon:
        app["icon"] = localize_asset(slug, app_slug, icon)
    app["screenshots"] = shots
    for src, dest in (
        ("ios", "ios"),
        ("google_play", "googlePlay"),
        ("android", "android"),
        ("user_guide", "userGuide"),
    ):
        val = get(src)
        if not val or val == "#":
            continue
        if src in FILE_FIELDS:
            val = localize_asset(slug, app_slug, val)
        app[dest] = val
    return app


def parse_sheet(ws: Worksheet) -> tuple[dict, list[str]]:
    warnings: list[str] = []
    header_row = find_header_row(ws)
    if not header_row:
        raise ValueError(f"{ws.title} 找不到表頭（需要 name 與 screenshot_1）")

    cols = col_index_map(ws, header_row)
    for key in COL_KEYS:
        if key not in cols:
            warnings.append(f"{ws.title}: 缺少欄 {key}")

    client_name = read_meta(ws, header_row, META_LABEL)
    if not client_name:
        warnings.append(f"{ws.title}: 未填 clientName（B1）")

    apps = []
    seen_slugs: set[str] = set()
    r = header_row + 1
    empty_streak = 0
    while r <= ws.max_row and empty_streak < 8:
        app = row_to_app(ws, r, cols, ws.title)
        if app is None:
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0
        display_name = app["name"]
        app_slug = name_to_slug(display_name)
        slug_err = validate_app_slug(ws.title, display_name, app_slug)
        if slug_err:
            warnings.append(slug_err)
        elif app_slug in seen_slugs:
            warnings.append(f"{ws.title}「{display_name}」的 slug「{app_slug}」與同表其他 APP 重複")
        else:
            seen_slugs.add(app_slug)
        if not app.get("screenshots"):
            warnings.append(f"{ws.title} 第 {r} 列「{app['name']}」沒有截圖")
        else:
            c1 = cols.get("screenshot_1")
            c2 = cols.get("screenshot_2")
            has1 = bool(c1 and cell_href(ws.cell(r, c1)))
            has2 = bool(c2 and cell_href(ws.cell(r, c2)))
            if has2 and not has1:
                warnings.append(
                    f"{ws.title} 第 {r} 列「{app['name']}」截圖寫在 screenshot_2；"
                    "請改放到 screenshot_1（單機只要第 1 張）"
                )
        if len(app.get("screenshots") or []) > 2:
            warnings.append(
                f"{ws.title} 第 {r} 列「{app['name']}」截圖超過 2 張，匯出只留前 2 張"
            )
            app["screenshots"] = app["screenshots"][:2]
        apps.append(app)
        r += 1

    if not apps:
        raise ValueError(f"{ws.title} 沒有任何 APP 列")

    payload = {"clientName": client_name, "apps": apps}
    qr_code = read_meta(ws, header_row, QR_LABEL)
    if qr_code:
        payload["qrCode"] = localize_qr(ws.title, qr_code)
    return payload, warnings


def resolve_file(slug: str, value: str) -> Path | None:
    if not value or value.startswith("#") or "://" in value:
        return None
    path = Path(value)
    if path.is_absolute():
        return path
    rooted = ROOT / value
    if rooted.exists():
        return rooted
    return CLIENTS_DIR / slug / value


def check_files(slug: str, payload: dict) -> list[str]:
    missing = []
    qr = payload.get("qrCode", "")
    if qr:
        path = resolve_file(slug, qr)
        if path is not None and not path.exists():
            missing.append(f"{slug}: 找不到 QR 圖 {qr}")
    for app in payload["apps"]:
        values = [app.get("icon", ""), app.get("userGuide", "")]
        values.extend(app.get("screenshots") or [])
        for value in values:
            path = resolve_file(slug, value)
            if path is None:
                continue
            if not path.exists():
                missing.append(f"{slug} / {app['name']}: 找不到 {value}")
    return missing


def write_json(slug: str, payload: dict) -> Path:
    folder = CLIENTS_DIR / slug
    folder.mkdir(parents=True, exist_ok=True)
    dest = folder / "client.json"
    dest.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return dest


def write_data_js(payload: dict, slug: str) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2)
    DATA_JS.write_text(
        "// Generated by scripts/clients_xlsx.py preview — 不要手改。\n"
        f"// 來源：clients.xlsx → {slug}\n"
        "window.CLIENT = "
        + body
        + ";\n",
        encoding="utf-8",
    )


def load_book() -> object:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"找不到 {XLSX_PATH.name}。請先執行 python3 scripts/clients_xlsx.py init"
        )
    return load_workbook(XLSX_PATH, data_only=True)


def ensure_xlsx_columns() -> None:
    if not XLSX_PATH.exists():
        return
    wb = load_workbook(XLSX_PATH)
    if ensure_book_columns(wb):
        wb.save(XLSX_PATH)
        print(f"已在 {XLSX_PATH.name} 補上缺少的欄（含 tagline）")


def export_all(check: bool) -> tuple[dict[str, dict], list[str], list[str]]:
    ensure_xlsx_columns()
    wb = load_book()
    payloads: dict[str, dict] = {}
    errors: list[str] = []
    warnings: list[str] = []

    for name in wb.sheetnames:
        err = validate_slug(name)
        if name.startswith(SKIP_PREFIX):
            continue
        if err:
            errors.append(err)
            continue
        try:
            payload, sheet_warnings = parse_sheet(wb[name])
        except ValueError as exc:
            errors.append(str(exc))
            continue
        warnings.extend(sheet_warnings)
        if check:
            warnings.extend(check_files(name, payload))
        payloads[name] = payload
    return payloads, errors, warnings


def cmd_export(check: bool) -> int:
    payloads, errors, warnings = export_all(check)
    for slug, payload in payloads.items():
        dest = write_json(slug, payload)
        n = len(payload["apps"])
        print(f"匯出 {slug}（{n} 款 APP）→ {dest.relative_to(ROOT)}")
    for msg in warnings:
        print(f"提示：{msg}")
    for msg in errors:
        sys.stderr.write(f"錯誤：{msg}\n")
    if not payloads and not errors:
        print("沒有可匯出的工作表。請複製 _template 並改成短名。")
        return 1
    return 1 if errors else 0


def cmd_preview(slug: str, check: bool) -> int:
    code = cmd_export(check)
    if code != 0:
        return code
    dest = CLIENTS_DIR / slug / "client.json"
    if not dest.exists():
        sys.stderr.write(f"沒有 {slug}。請確認工作表名稱，或先 export。\n")
        return 1
    payload = json.loads(dest.read_text(encoding="utf-8"))
    write_data_js(payload, slug)
    print(f"預覽已寫入 {DATA_JS.relative_to(ROOT)}（{slug}）。重新整理本機頁面即可。")
    return 0


def owns_path(slug: str, path: Path) -> bool:
    try:
        path.resolve().relative_to((CLIENTS_DIR / slug).resolve())
        return True
    except ValueError:
        return False


def in_other_product(slug: str, dest_dir: Path, path: Path) -> bool:
    client_dir = (CLIENTS_DIR / slug).resolve()
    dest = dest_dir.resolve()
    resolved = path.resolve()
    try:
        rel = resolved.relative_to(client_dir)
    except ValueError:
        return True
    if len(rel.parts) == 1:
        return False
    other = (client_dir / rel.parts[0])
    return other.is_dir() and other != dest


def unique_files(paths: list[Path]) -> list[Path]:
    seen: set[Path] = set()
    out: list[Path] = []
    for path in paths:
        if not path.exists() or not path.is_file():
            continue
        key = path.resolve()
        if key in seen:
            continue
        seen.add(key)
        out.append(path)
    return out


def find_client_file(
    slug: str, app_slug: str, display_name: str, value: str, stem: str
) -> Path | None:
    if is_remote_or_shared(value):
        return None
    prefix = f"clients/{slug}/"
    if value.startswith("clients/") and not value.startswith(prefix):
        return None
    client_dir = CLIENTS_DIR / slug
    dest_dir = client_dir / app_slug
    legacy_dir = client_dir / legacy_product_folder(display_name)
    basename = Path(value).name
    search_dirs = [dest_dir]
    if legacy_dir != dest_dir:
        search_dirs.append(legacy_dir)
    hits: list[Path] = []
    for folder in search_dirs:
        hits.append(folder / basename)
        if folder.exists() and stem:
            hits.extend(p for p in folder.iterdir() if p.is_file() and p.stem == stem)
    if client_dir.exists():
        hits.append(client_dir / basename)
        hits.extend(
            p
            for p in client_dir.iterdir()
            if p.is_file() and (p.name == basename or (stem and p.stem == stem))
        )
        for folder in search_dirs:
            if not folder.exists():
                continue
            for path in folder.rglob(basename):
                if path.is_file() and not in_other_product(slug, dest_dir, path):
                    hits.append(path)
    found = [p for p in unique_files(hits) if owns_path(slug, p)]
    in_dest_canon = [p for p in found if p.parent.resolve() == dest_dir.resolve() and p.stem == stem]
    if in_dest_canon:
        return in_dest_canon[0]
    in_dest = [p for p in found if p.parent.resolve() == dest_dir.resolve()]
    if in_dest:
        return in_dest[0]
    in_legacy = [p for p in found if legacy_dir != dest_dir and p.parent.resolve() == legacy_dir.resolve()]
    if in_legacy:
        return in_legacy[0]
    loose = [p for p in found if p.parent.resolve() == client_dir.resolve()]
    if loose:
        return loose[0]
    return found[0] if found else None


def move_to_canonical(src: Path, dest: Path) -> str:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if src.resolve() == dest.resolve():
        return "same"
    if dest.exists():
        return "conflict"
    shutil.move(str(src), str(dest))
    return "moved"


def migrate_product_folder(slug: str, display_name: str, app_slug: str) -> list[str]:
    notes: list[str] = []
    client_dir = CLIENTS_DIR / slug
    legacy_name = legacy_product_folder(display_name)
    if legacy_name == app_slug:
        return notes
    old_dir = client_dir / legacy_name
    new_dir = client_dir / app_slug
    if not old_dir.exists():
        return notes
    if legacy_name.lower() == app_slug.lower():
        temp_dir = client_dir / f".{app_slug}-slug-tmp"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        shutil.move(str(old_dir), str(temp_dir))
        shutil.move(str(temp_dir), str(new_dir))
        notes.append(f"資料夾 {legacy_name} → {app_slug}（大小寫）")
        return notes
    if not new_dir.exists():
        shutil.move(str(old_dir), str(new_dir))
        notes.append(f"資料夾 {old_dir.relative_to(ROOT)} → {new_dir.relative_to(ROOT)}")
        return notes
    for path in sorted(old_dir.rglob("*")):
        if not path.is_file() or path.name == ".DS_Store":
            continue
        dest = new_dir / path.relative_to(old_dir)
        dest.parent.mkdir(parents=True, exist_ok=True)
        if dest.exists():
            continue
        shutil.move(str(path), str(dest))
        notes.append(f"{path.relative_to(ROOT)} → {dest.relative_to(ROOT)}")
    for junk in old_dir.rglob(".DS_Store"):
        junk.unlink()
    try:
        old_dir.rmdir()
        notes.append(f"刪空資料夾 {old_dir.relative_to(ROOT)}")
    except OSError:
        pass
    return notes


def prune_empty_product_dirs(slug: str) -> list[str]:
    notes: list[str] = []
    client_dir = CLIENTS_DIR / slug
    if not client_dir.exists():
        return notes
    for folder in sorted(client_dir.iterdir(), reverse=True):
        if not folder.is_dir():
            continue
        leftover = [p for p in folder.rglob("*") if p.is_file() and p.name != ".DS_Store"]
        if leftover:
            continue
        for junk in folder.rglob(".DS_Store"):
            junk.unlink()
        try:
            folder.rmdir()
            notes.append(f"刪空資料夾 {folder.relative_to(ROOT)}")
        except OSError:
            pass
    return notes


def cmd_sync() -> int:
    if not XLSX_PATH.exists():
        sys.stderr.write(f"找不到 {XLSX_PATH.name}。\n")
        return 1
    wb = load_workbook(XLSX_PATH)
    notes: list[str] = []
    errors: list[str] = []
    excel_changed = ensure_book_columns(wb)
    if "_readme" in wb.sheetnames:
        write_readme(wb["_readme"])
    else:
        wb.create_sheet("_readme", 0)
        write_readme(wb["_readme"])
    excel_changed = True
    rewrites: list[tuple[str, str]] = []

    for name in wb.sheetnames:
        if name.startswith(SKIP_PREFIX):
            continue
        err = validate_slug(name)
        if err:
            errors.append(err)
            continue
        ws = wb[name]
        header_row = find_header_row(ws)
        if not header_row:
            errors.append(f"{name} 找不到表頭")
            continue
        cols = col_index_map(ws, header_row)
        qr_raw = read_meta(ws, header_row, QR_LABEL)
        if qr_raw and not is_remote_or_shared(qr_raw):
            prefix = f"clients/{name}/"
            if not (qr_raw.startswith("clients/") and not qr_raw.startswith(prefix)):
                src = find_client_file(name, "", "", qr_raw, QR_STEM)
                if src is None:
                    src = unique_files(
                        [
                            CLIENTS_DIR / name / Path(qr_raw).name,
                            *((CLIENTS_DIR / name).rglob(Path(qr_raw).name)
                              if (CLIENTS_DIR / name).exists()
                              else []),
                        ]
                    )
                    src = src[0] if src else None
                if src is None:
                    notes.append(f"{name}: 找不到 QR {qr_raw}")
                else:
                    dest = CLIENTS_DIR / name / f"{QR_STEM}{src.suffix}"
                    result = move_to_canonical(src, dest)
                    if result == "moved":
                        notes.append(f"{name}: QR {src.name} → {dest.relative_to(ROOT)}")
                        rewrites.append(
                            (
                                src.relative_to(ROOT).as_posix(),
                                dest.relative_to(ROOT).as_posix(),
                            )
                        )
                    elif result == "conflict":
                        errors.append(f"{name}: QR 目標已存在 {dest.relative_to(ROOT)}")
                    if result != "conflict":
                        for r in range(1, header_row):
                            if cell_str(ws.cell(r, 1).value) == QR_LABEL:
                                if cell_str(ws.cell(r, 2).value) != dest.name:
                                    ws.cell(r, 2).value = dest.name
                                    excel_changed = True
                                break

        r = header_row + 1
        empty_streak = 0
        seen_slugs: set[str] = set()
        while r <= ws.max_row and empty_streak < 8:
            name_col = cols.get("name")
            product = cell_href(ws.cell(r, name_col)) if name_col else ""
            if not product:
                empty_streak += 1
                r += 1
                continue
            empty_streak = 0
            app_slug = name_to_slug(product)
            slug_err = validate_app_slug(name, product, app_slug)
            if slug_err:
                errors.append(slug_err)
                r += 1
                continue
            if app_slug in seen_slugs:
                errors.append(f"{name}「{product}」的 slug「{app_slug}」與同表其他 APP 重複")
                r += 1
                continue
            seen_slugs.add(app_slug)
            notes.extend(migrate_product_folder(name, product, app_slug))
            dest_dir = CLIENTS_DIR / name / app_slug
            for field, stem in CANON_STEMS.items():
                col = cols.get(field)
                if not col:
                    continue
                raw = cell_href(ws.cell(r, col))
                if not raw or is_remote_or_shared(raw):
                    continue
                src = find_client_file(name, app_slug, product, raw, stem)
                if src is None:
                    notes.append(f"{name} / {product}: 找不到 {field} {raw}")
                    continue
                dest = dest_dir / f"{stem}{src.suffix}"
                result = move_to_canonical(src, dest)
                if result == "moved":
                    notes.append(
                        f"{name} / {product}: {src.name} → {dest.relative_to(ROOT)}"
                    )
                    rewrites.append(
                        (
                            src.relative_to(ROOT).as_posix(),
                            dest.relative_to(ROOT).as_posix(),
                        )
                    )
                elif result == "conflict":
                    errors.append(
                        f"{name} / {product}: 目標已存在 {dest.relative_to(ROOT)}"
                    )
                    continue
                if cell_str(ws.cell(r, col).value) != dest.name:
                    ws.cell(r, col).value = dest.name
                    excel_changed = True
            r += 1
        notes.extend(prune_empty_product_dirs(name))

    if rewrites:
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith(SKIP_PREFIX):
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    text = cell_str(cell.value)
                    for old, new in rewrites:
                        if text == old:
                            cell.value = new
                            excel_changed = True
                            notes.append(f"{sheet_name}: 路徑 {old} → {new}")
                            break

    if excel_changed:
        wb.save(XLSX_PATH)
        print(f"已更新 {XLSX_PATH.name} 檔名")
    for msg in notes:
        print(msg)
    for msg in errors:
        sys.stderr.write(f"錯誤：{msg}\n")
    if errors:
        return 1
    return cmd_export(True)


def iter_pack_files(slug: str) -> list[Path]:
    folder = CLIENTS_DIR / slug
    if not folder.exists():
        return []
    files = []
    for path in sorted(folder.rglob("*")):
        if not path.is_file() or path.name == ".DS_Store":
            continue
        rel = path.relative_to(folder)
        if len(rel.parts) == 1:
            if rel.name == "client.json" or path.stem == QR_STEM:
                files.append(path)
            continue
        files.append(path)
    return files


def cmd_pack() -> int:
    code = cmd_export(True)
    if code != 0:
        return code
    payloads, errors, _warnings = export_all(False)
    if errors:
        return 1
    slugs = [s for s in payloads if s not in PACK_SKIP_SLUGS]
    dest = PACK_ZIP
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists():
        dest.unlink()
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in PACK_FILES:
            path = ROOT / name
            if path.exists():
                zf.write(path, name)
        for folder in PACK_DIRS:
            base = ROOT / folder
            if not base.exists():
                continue
            for path in sorted(base.rglob("*")):
                if not path.is_file() or path.name == ".DS_Store":
                    continue
                zf.write(path, path.relative_to(ROOT).as_posix())
        for slug in slugs:
            for path in iter_pack_files(slug):
                zf.write(path, path.relative_to(ROOT).as_posix())
    print(f"打包 {dest.relative_to(ROOT)}")
    print("客戶：" + ", ".join(slugs) if slugs else "客戶：無")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="以 Excel 管理客戶資料（一 sheet 一間公司）"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="產生 clients.xlsx 模板（含範例）")
    p_init.add_argument("--force", action="store_true", help="覆蓋已存在的 clients.xlsx")

    p_export = sub.add_parser("export", help="匯出全部客戶 sheet 成 client.json")
    p_export.add_argument(
        "--check",
        action="store_true",
        help="檢查圖檔／說明書路徑是否存在（僅提示）",
    )

    p_preview = sub.add_parser("preview", help="匯出後把指定客戶寫進 js/data.js")
    p_preview.add_argument("slug", help="工作表名稱，例如 example-pd")
    p_preview.add_argument("--check", action="store_true")

    sub.add_parser("sync", help="依命名規範改檔名、搬進產品資料夾，並更新 Excel 後匯出")
    sub.add_parser("pack", help="匯出後打上線 zip 到 dist/csi-appstore.zip")

    args = parser.parse_args()
    if args.cmd == "init":
        return cmd_init(args.force)
    if args.cmd == "export":
        return cmd_export(args.check)
    if args.cmd == "preview":
        return cmd_preview(args.slug, args.check)
    if args.cmd == "sync":
        return cmd_sync()
    if args.cmd == "pack":
        return cmd_pack()
    return 1


if __name__ == "__main__":
    sys.exit(main())
